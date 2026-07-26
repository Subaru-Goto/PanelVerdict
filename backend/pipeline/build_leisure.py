"""Stage 1 — build a country's leisure profile table from its time-use survey.

One keyless source per country, mirroring `build_oecd`. Germany comes from
Eurostat's HETUS 2020 round (`tus_20age`) as JSON-stat; Japan from
社会生活基本調査 2021 table 1-1, which e-Stat serves as a spreadsheet without an
application ID (its CSV and JSON endpoints need one).

Only participation rates are read: the pool needs to know *whether* someone is
into something, not for how long. Durations would have to be rendered as text
("175 minutes a day") that is not comparable between surveys, and would need a
spread model no survey supports.

Every emitted row cites the table and activities it came from, and anything
combined rather than read straight off is declared in `imputations` and
committed beside the data. Design: issues/006i.
"""

import csv
import json
import re
import urllib.request
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.schemas import LeisureCategory, Locale

_Gender = Literal["male", "female"]
_GENDERS: tuple[_Gender, ...] = ("male", "female")

_AgeBand = Literal["15-24", "25-34", "35-44", "45-54", "55-64", "65+"]
_AGE_BANDS: tuple[_AgeBand, ...] = (
    "15-24",
    "25-34",
    "35-44",
    "45-54",
    "55-64",
    "65+",
)

# Everything `openpyxl.iter_rows(values_only=True)` can hand back.
_CellValue = str | int | float | bool | datetime | date | time | timedelta | None

# Germany: Eurostat acl18 activity codes. Several categories take more than one
# code because Eurostat splits finer than Japan's diary does, and the coarser
# survey sets the granularity — print media joins TV and radio, and walking
# joins sport, because Japan publishes them that way and a category has to mean
# the same thing in both tables.
_EUROSTAT_CODES: dict[LeisureCategory, tuple[str, ...]] = {
    LeisureCategory.TV_MEDIA: ("AC821", "AC831", "AC81_X_812"),
    LeisureCategory.SOCIALIZING: ("AC511", "AC512_513_519", "AC514-516"),
    LeisureCategory.SPORTS_EXERCISE: ("AC6_X_611", "AC611"),
    LeisureCategory.VOLUNTEERING: ("AC4",),
    LeisureCategory.HOBBIES_AND_GAMES: (
        "AC711_712_719_731_732_739",
        "AC72",
        "AC733-735",
        "AC812",
    ),
}

_EUROSTAT_AGE: dict[_AgeBand, tuple[str, ...]] = {
    "15-24": ("Y15-24",),
    "25-34": ("Y25-34",),
    "35-44": ("Y35-44",),
    "45-54": ("Y45-54",),
    "55-64": ("Y55-64",),
    "65+": ("Y_GE65",),
}

_EUROSTAT_BASE = (
    "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/tus_20age"
)
_EUROSTAT_CITATION = "eurostat:tus_20age"
_EUROSTAT_SEX: dict[str, str] = {"M": "male", "F": "female", "T": "total"}
_EUROSTAT_AGE_TOTAL = "TOTAL"
_EUROSTAT_RATE_UNIT = "PTP_RT"

# Dimensions we index across; every other dimension must be pinned to one value
# by the query, or the flat value index below would silently mis-map.
_INDEXED_DIMS = ("sex", "age", "acl18", "unit")

# Japan: keyed on the sheet's own English headers, so nothing here depends on
# matching Japanese text. Each category is one diary activity — the taxonomy is
# coarser than the pool's, never finer, so Japan never combines activities.
_ESTAT_ACTIVITIES: dict[LeisureCategory, tuple[str, ...]] = {
    LeisureCategory.TV_MEDIA: (
        "Watching TV, listening to the radio, reading newspapers or magazines",
    ),
    LeisureCategory.SOCIALIZING: ("Social life",),
    LeisureCategory.SPORTS_EXERCISE: ("Sports",),
    LeisureCategory.VOLUNTEERING: ("Volunteer and social activities",),
    LeisureCategory.HOBBIES_AND_GAMES: ("Hobbies and amusements",),
}

# Japan publishes 15-24, 45-54, 55-64 and 65+ ready-made; the two middle bands
# have to be combined from its five-year groups, which is why the parser also
# collects each group's population.
_ESTAT_AGE: dict[_AgeBand, tuple[str, ...]] = {
    "15-24": ("15~24歳",),
    "25-34": ("25~29歳", "30~34歳"),
    "35-44": ("35~39歳", "40~44歳"),
    "45-54": ("45~54歳",),
    "55-64": ("55~64歳",),
    "65+": ("65歳以上",),
}

_ESTAT_URL = (
    "https://www.e-stat.go.jp/stat-search/file-download"
    "?statInfId=000032262854&fileKind=0"
)
_ESTAT_CITATION = "e-stat:shakai-seikatsu-2021:table1-1"
_ESTAT_SEX: dict[str, str] = {"男": "male", "女": "female", "総数": "total"}
# The all-ages row of each block simply leaves the age cell blank.
_ESTAT_AGE_TOTAL = ""
_ESTAT_RATE_LABEL = "行動者率"
_ESTAT_OTHER_BLOCKS = ("総平均時間", "行動者平均時間")
_ESTAT_POPULATION_HEADER = "Population 10 years and over (1000)"

_LEISURE_COLUMNS = ["category", "gender", "age_band", "participation_rate", "source"]


class LeisureRow(BaseModel):
    """One (category, gender, age band) cell of a country's leisure table."""

    category: LeisureCategory
    gender: _Gender
    age_band: _AgeBand
    participation_rate: float = Field(gt=0, le=1)
    source: str


class LeisureBuildResult(BaseModel):
    """A country's built leisure table and the fidelity it was built at."""

    country: Locale
    rows: list[LeisureRow]
    imputations: list[str]


@dataclass(frozen=True)
class SurveyCells:
    """A survey's published participation rates, keyed by its own gender label,
    age label and activity — plus the population of each age group, where the
    survey gives one, so its groups can be combined into our bands.

    `totals` names the survey's own "all genders" and "all ages" labels. Not
    every survey crosses both: ATUS publishes participation by sex but not by
    age. Keeping the totals lets a missing cell fall back to the coarsest one
    that exists rather than dropping the country.
    """

    rates: dict[tuple[str, str, str], float]
    populations: dict[tuple[str, str], float]
    totals: tuple[str, str]


def _eurostat_url(country: Locale) -> str:
    return f"{_EUROSTAT_BASE}?format=JSON&lang=en&geo={country.value}"


def parse_jsonstat(text: str) -> SurveyCells:
    """Flatten a JSON-stat payload into participation rates.

    JSON-stat stores values against a single row-major index over all
    dimensions, so strides are derived from `size` rather than assumed — a
    dimension reordering upstream would otherwise mis-map every cell silently.
    The unit is read from the data rather than filtered in the query, because
    Eurostat accepts a `unit=` parameter and then ignores it.
    """
    payload = json.loads(text)
    dims: list[str] = payload["id"]
    sizes: list[int] = payload["size"]
    strides: dict[str, int] = {}
    stride = 1
    for dim, size in zip(reversed(dims), reversed(sizes)):
        strides[dim] = stride
        stride *= size
    unpinned = [
        dim for dim, size in zip(dims, sizes) if size > 1 and dim not in _INDEXED_DIMS
    ]
    if unpinned:
        raise ValueError(f"query must pin one value per dimension; got {unpinned}")

    index = {
        dim: payload["dimension"][dim]["category"]["index"] for dim in _INDEXED_DIMS
    }
    rate_index = index["unit"].get(_EUROSTAT_RATE_UNIT)
    if rate_index is None:
        raise ValueError(f"payload publishes no {_EUROSTAT_RATE_UNIT} unit")

    values = payload["value"]
    rates: dict[tuple[str, str, str], float] = {}
    for sex, sex_i in index["sex"].items():
        gender = _EUROSTAT_SEX.get(sex)
        if gender is None:
            continue
        for age, age_i in index["age"].items():  # includes the survey's own totals
            for activity, activity_i in index["acl18"].items():
                flat = (
                    sex_i * strides["sex"]
                    + age_i * strides["age"]
                    + activity_i * strides["acl18"]
                    + rate_index * strides["unit"]
                )
                raw = values.get(str(flat))
                if raw is not None:
                    rates[(gender, age, activity)] = float(raw) / 100
    return SurveyCells(
        rates=rates, populations={}, totals=(_EUROSTAT_SEX["T"], _EUROSTAT_AGE_TOTAL)
    )


def _cell_text(value: _CellValue) -> str:
    """Header cells carry line breaks, so collapse whitespace before comparing
    against the published labels."""
    return " ".join(str(value).split()) if value is not None else ""


def _estat_age_label(value: _CellValue) -> str:
    """Age labels carry trailing furigana and a fullwidth tilde; fold both so
    they compare against the band names as written."""
    text = _cell_text(value).replace("〜", "~").replace("～", "~")
    return re.sub(r"[ァ-ヶー]+$", "", text)


def _estat_grid(payload: bytes) -> list[list[_CellValue]]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        return [
            list(row)
            for row in workbook[workbook.sheetnames[0]].iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


def _estat_rate_columns(grid: list[list[_CellValue]]) -> dict[str, int]:
    """Locate each wanted activity's column within the 行動者率 block.

    The sheet lays three metric blocks side by side over one repeated set of
    activity headers, so a column's meaning comes from the block label above it.
    Read positionally this would return a duration where a percentage belongs —
    the same silent failure Eurostat's ignored `unit` filter produces.
    """
    blocks: dict[int, str] = {}
    for row in grid:
        labelled = {
            col: label
            for col, value in enumerate(row)
            for label in (_ESTAT_RATE_LABEL, *_ESTAT_OTHER_BLOCKS)
            if _cell_text(value).startswith(label)
        }
        # Distinct labels, not just three hits: the sheet's title line repeats
        # every block name in one cell and would otherwise pass as the header.
        if len(labelled) == len(set(labelled.values())) == len(_ESTAT_OTHER_BLOCKS) + 1:
            blocks = labelled
            break
    if not blocks:
        raise ValueError("no row carries all three metric-block labels")
    rate_start = next(
        col for col, label in blocks.items() if label == _ESTAT_RATE_LABEL
    )
    block_end = min(
        (col for col in blocks if col > rate_start), default=max(len(r) for r in grid)
    )

    wanted = {name for names in _ESTAT_ACTIVITIES.values() for name in names}
    columns = {
        _cell_text(value): col
        for row in grid
        for col, value in enumerate(row)
        if rate_start <= col < block_end and _cell_text(value) in wanted
    }
    missing = wanted - set(columns)
    if missing:
        raise ValueError(f"activity headers absent from the rate block: {missing}")
    return columns


def parse_estat_table(payload: bytes) -> SurveyCells:
    """Flatten e-Stat table 1-1 into participation rates by gender and age group.

    Nothing is addressed by row or column number: the metric block comes from
    its label, activities from the sheet's own English headers, and the age rows
    from the sex label that starts each block.
    """
    grid = _estat_grid(payload)
    columns = _estat_rate_columns(grid)

    sex_col = next(
        (
            col
            for row in grid
            for col, value in enumerate(row)
            if _cell_text(value) in _ESTAT_SEX
        ),
        None,
    )
    if sex_col is None:
        raise ValueError(f"no column carries a bare sex label ({'/'.join(_ESTAT_SEX)})")
    population_col = next(
        (
            col
            for row in grid
            for col, value in enumerate(row)
            if _cell_text(value) == _ESTAT_POPULATION_HEADER
        ),
        None,
    )
    if population_col is None:
        raise ValueError(f"no column headed {_ESTAT_POPULATION_HEADER!r}")

    wanted_ages = {label for labels in _ESTAT_AGE.values() for label in labels}
    wanted_ages.add(_ESTAT_AGE_TOTAL)
    rates: dict[tuple[str, str, str], float] = {}
    populations: dict[tuple[str, str], float] = {}
    gender: str | None = None
    for row in grid:
        gender = _ESTAT_SEX.get(_cell_text(row[sex_col]), gender)
        age = _estat_age_label(row[sex_col + 1])
        population = row[population_col]
        # A numeric population is what separates a data row from the headers,
        # spacers and (再掲) labels that also leave the age cell blank.
        if gender is None or age not in wanted_ages:
            continue
        if not isinstance(population, int | float):
            continue
        for activity, col in columns.items():
            value = row[col]
            if not isinstance(value, int | float):
                raise ValueError(f"{gender} {age} {activity!r} reads {value!r}")
            rates[(gender, age, activity)] = value / 100
        populations[(gender, age)] = float(population)
    return SurveyCells(
        rates=rates,
        populations=populations,
        totals=(_ESTAT_SEX["総数"], _ESTAT_AGE_TOTAL),
    )


# Coarsest-first is wrong, so these are ordered finest-first: the first key that
# the survey actually publishes wins.
_Conditioning = Literal["gender and age", "gender only", "age only", "neither"]


def _cell_rate(
    cells: SurveyCells, gender: str, age: str, activity: str
) -> tuple[float, _Conditioning]:
    """The finest published rate for one cell, and what it was conditioned on.

    Surveys differ in what they cross: Eurostat and e-Stat publish activity by
    gender and age, ATUS by gender alone. Rather than drop a country or leave
    holes, a missing cell falls back to the coarsest published prior — the
    gender's own all-ages figure, then the all-gender band, then the national
    total — and the builder declares which was used.
    """
    gender_total, age_total = cells.totals
    candidates: tuple[tuple[str, str, _Conditioning], ...] = (
        (gender, age, "gender and age"),
        (gender, age_total, "gender only"),
        (gender_total, age, "age only"),
        (gender_total, age_total, "neither"),
    )
    for gender_key, age_key, conditioning in candidates:
        rate = cells.rates.get((gender_key, age_key, activity))
        if rate is not None:
            return rate, conditioning
    raise KeyError(f"no published rate for {activity!r} at any conditioning level")


def _band_rate(
    cells: SurveyCells,
    gender: str,
    age_labels: tuple[str, ...],
    activities: tuple[str, ...],
) -> tuple[float, set[_Conditioning]]:
    """One band's participation rate for one category, and the conditioning
    levels that went into it.

    Two different combinations, and they are not interchangeable. Across
    activities the rate is the union under independence, because one person
    appears under several activities and rates cannot be added. Across the
    survey's own age groups it is the population-weighted mean, because those
    groups are disjoint sets of people.
    """
    if len(age_labels) > 1 and not cells.populations:
        raise ValueError(
            f"{age_labels} must be combined but the survey publishes no populations"
        )
    levels: set[_Conditioning] = set()
    weighted = 0.0
    total = 0.0
    for label in age_labels:
        not_participating = 1.0
        for activity in activities:
            rate, conditioning = _cell_rate(cells, gender, label, activity)
            levels.add(conditioning)
            not_participating *= 1 - rate
        weight = cells.populations.get((gender, label), 1.0)
        weighted += (1 - not_participating) * weight
        total += weight
    return weighted / total, levels


def _leisure_rows(
    cells: SurveyCells,
    activities: dict[LeisureCategory, tuple[str, ...]],
    ages: dict[_AgeBand, tuple[str, ...]],
    citation: str,
) -> tuple[list[LeisureRow], set[_Conditioning]]:
    """Every (category, gender, band) row, plus the conditioning levels used."""
    rows: list[LeisureRow] = []
    levels: set[_Conditioning] = set()
    for category, keys in activities.items():
        for gender in _GENDERS:
            for band in _AGE_BANDS:
                rate, used = _band_rate(cells, gender, ages[band], keys)
                levels |= used
                rows.append(
                    LeisureRow(
                        category=category,
                        gender=gender,
                        age_band=band,
                        participation_rate=rate,
                        source=f"{citation}:{'+'.join(keys)}",
                    )
                )
    return rows, levels


def _conditioning_note(levels: set[_Conditioning]) -> str | None:
    """Declare any cell that fell back to a coarser prior than gender x age."""
    fallbacks = sorted(level for level in levels if level != "gender and age")
    if not fallbacks:
        return None
    return (
        f"some cells are not conditioned on both gender and age ({', '.join(fallbacks)})"
        ": the survey publishes no finer breakdown, so the coarser published rate "
        "stands in as the prior for every band it covers"
    )


def _build_germany(fetch: Callable[[str], bytes]) -> LeisureBuildResult:
    cells = parse_jsonstat(fetch(_eurostat_url(Locale.DE)).decode())
    rows, levels = _leisure_rows(
        cells, _EUROSTAT_CODES, _EUROSTAT_AGE, _EUROSTAT_CITATION
    )
    combined = [
        category.value for category, codes in _EUROSTAT_CODES.items() if len(codes) > 1
    ]
    imputations = [
        f"categories built from several acl18 codes ({', '.join(combined)}) take "
        "the union under independence, 1-prod(1-r), since the same person is "
        "counted under each code and published rates cannot be added",
        "every age band maps to one published Eurostat band, so no age group "
        "is combined",
    ]
    note = _conditioning_note(levels)
    return LeisureBuildResult(
        country=Locale.DE,
        rows=rows,
        imputations=imputations + ([note] if note else []),
    )


def _build_japan(fetch: Callable[[str], bytes]) -> LeisureBuildResult:
    cells = parse_estat_table(fetch(_ESTAT_URL))
    rows, levels = _leisure_rows(cells, _ESTAT_ACTIVITIES, _ESTAT_AGE, _ESTAT_CITATION)
    combined = [band for band, labels in _ESTAT_AGE.items() if len(labels) > 1]
    note = _conditioning_note(levels)
    return LeisureBuildResult(
        country=Locale.JP,
        rows=rows,
        imputations=[
            f"the {' and '.join(combined)} bands are not published ready-made and "
            "are the population-weighted mean of the survey's five-year groups, "
            "using the populations the table publishes beside them",
            "rates are diary-day rates (主行動, primary activity only), far below "
            "the past-year 行動者率 the survey also publishes — volunteering is "
            "1.8% of men on a given day against ~26% in a year",
            "socializing is 交際・付き合い, which covers visits and ceremonies but "
            "not conversation at home; Eurostat's socializing includes it, so the "
            "two countries' figures are not comparable to each other",
            *([note] if note else []),
        ],
    )


def build_leisure(
    country: Locale, *, fetch: Callable[[str], bytes]
) -> LeisureBuildResult:
    """Build one country's leisure table: a participation rate per category,
    gender and age band."""
    if country is Locale.DE:
        return _build_germany(fetch)
    if country is Locale.JP:
        return _build_japan(fetch)
    raise NotImplementedError(
        "US needs ATUS Table A-1, published only as PDF; bls.gov also serves no "
        "participation rates by age — slice 1b, see issues/006i"
    )


def write_leisure(result: LeisureBuildResult, dest_dir: Path) -> None:
    """Write the committed per-country CSV plus its provenance sidecar.

    The sidecar matters: a caveat that only reaches stdout is invisible to
    anyone who later reads the CSV (`build_oecd.write_joint` sets the precedent).
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    stem = result.country.value.lower()
    with (dest_dir / f"{stem}.csv").open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_LEISURE_COLUMNS)
        writer.writeheader()
        for row in result.rows:
            writer.writerow(
                {
                    "category": row.category.value,
                    "gender": row.gender,
                    "age_band": row.age_band,
                    "participation_rate": f"{row.participation_rate:.4f}",
                    "source": row.source,
                }
            )
    (dest_dir / f"{stem}.meta.json").write_text(
        json.dumps(
            {"country": result.country.value, "imputations": result.imputations},
            indent=2,
        )
        + "\n"
    )


def _http_fetch(url: str) -> bytes:
    request = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (panelverdict pipeline)"}
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


if __name__ == "__main__":
    import sys

    target = Locale(sys.argv[1])
    dest = Path(__file__).parents[1] / "app" / "data" / "leisure"
    built = build_leisure(target, fetch=_http_fetch)
    write_leisure(built, dest)
    print(f"wrote {target.value}: {len(built.rows)} rows")
    for note in built.imputations:
        print(f"  declared: {note}")
